import sys
import re
import openpyxl
from openpyxl.styles import Font as OpenpyxlFont, PatternFill
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QTableWidget, QPushButton, QHBoxLayout,
    QTableWidgetItem, QFileDialog, QMessageBox, QColorDialog,
    QFontDialog, QLineEdit, QTabWidget, QLabel, QMenuBar, QMenu, QHeaderView
)
from PyQt5.QtGui import QColor, QFont
from PyQt5.QtCore import Qt, pyqtSignal


class CustomTableWidget(QTableWidget):
    # Sinal personalizado para deletar células via teclado
    deletePressed = pyqtSignal()

    def __init__(self, rows, columns, parent=None):
        super().__init__(rows, columns, parent)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Delete:
            self.deletePressed.emit()
        else:
            super().keyPressEvent(event)


class SimpleExcelWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.currently_editing = False          # Inicializa a flag de edição
        self.dependencies = {}                  # Inicializa dependências
        self.dependents = {}                    # Inicializa dependentes
        self.setup_ui()                         # Configura a interface do usuário

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Menu
        self.menu_bar = QMenuBar(self)
        layout.setMenuBar(self.menu_bar)
        self.setup_menu()

        # Botões de controle (no topo)
        self.setup_buttons()

        # Tab Widget
        self.tabs = QTabWidget(self)
        layout.addWidget(self.tabs)

        # Aba principal
        self.main_widget = QWidget(self)
        self.main_layout = QVBoxLayout(self.main_widget)
        self.tabs.addTab(self.main_widget, "Planilha")

        # Barra de fórmula (opcional)
        self.formula_bar = QLineEdit(self)
        self.formula_bar.setPlaceholderText("Digite uma fórmula ou valor aqui")
        self.formula_bar.returnPressed.connect(self.apply_formula)
        self.main_layout.addWidget(self.formula_bar)

        # Tabela
        self.table = CustomTableWidget(10, 5, self)
        self.table.setHorizontalHeaderLabels([chr(65 + i) for i in range(5)])
        self.table.itemSelectionChanged.connect(self.update_formula_bar)
        self.table.itemChanged.connect(self.handle_cell_edit)
        self.table.deletePressed.connect(self.delete_cell)  # Conectar o sinal personalizado
        self.table.cellDoubleClicked.connect(self.handle_cell_double_clicked)  # Conectar duplo clique
        self.main_layout.addWidget(self.table)

        # Configuração do Redimensionamento Automático das Colunas
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        # Definir Tamanho Mínimo e Padrão das Colunas
        self.table.horizontalHeader().setMinimumSectionSize(100)  # Define o tamanho mínimo (em pixels)
        self.table.horizontalHeader().setDefaultSectionSize(100)   # Define o tamanho padrão inicial (em pixels)

        # Configuração do Redimensionamento Automático das Linhas
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        # Definir Tamanho Mínimo e Padrão das Linhas
        self.table.verticalHeader().setMinimumSectionSize(30)  # Define o tamanho mínimo (em pixels)
        self.table.verticalHeader().setDefaultSectionSize(30)   # Define o tamanho padrão inicial (em pixels)

        # Inicializar todas as células com formatação padrão
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = QTableWidgetItem("")
                item.setBackground(QColor("#FFFFFF"))  # Fundo branco
                item.setForeground(QColor("#000000"))  # Texto preto
                font = QFont("Arial", 10)
                item.setFont(font)
                self.table.setItem(row, col, item)

        # Aba de ajuda
        self.setup_help_tab()

    def setup_menu(self):
        format_menu = QMenu("Formatar", self)
        self.menu_bar.addMenu(format_menu)
        # Atualmente, apenas o menu "Formatar" está presente sem ações adicionais

    def setup_buttons(self):
        button_layout = QHBoxLayout()

        add_row_btn = self.create_button("Adicionar Linha", self.add_row)
        remove_row_btn = self.create_button("Remover Linha", self.remove_row)
        add_col_btn = self.create_button("Adicionar Coluna", self.add_column)
        remove_col_btn = self.create_button("Remover Coluna", self.remove_column)
        save_excel_btn = self.create_button("Salvar (Excel)", self.save_excel)
        load_excel_btn = self.create_button("Carregar (Excel)", self.load_excel)
        format_bg_btn = self.create_button("Cor de Fundo", self.format_background)
        format_font_btn = self.create_button("Fonte", self.format_font)
        format_text_color_btn = self.create_button("Cor do Texto", self.format_text_color)
        delete_btn = self.create_button("Deletar Célula", self.delete_cell_button)  # Botão Delete separado

        button_layout.addWidget(add_row_btn)
        button_layout.addWidget(remove_row_btn)
        button_layout.addWidget(add_col_btn)
        button_layout.addWidget(remove_col_btn)
        button_layout.addWidget(save_excel_btn)
        button_layout.addWidget(load_excel_btn)
        button_layout.addWidget(format_bg_btn)
        button_layout.addWidget(format_font_btn)
        button_layout.addWidget(format_text_color_btn)
        button_layout.addWidget(delete_btn)  # Adicionado botão Delete à interface

        layout = self.layout()
        layout.insertLayout(0, button_layout)  # Inserir os botões no topo

    def create_button(self, text, slot):
        button = QPushButton(text)
        button.clicked.connect(slot)
        return button

    def setup_help_tab(self):
        self.help_widget = QWidget(self)
        self.help_layout = QVBoxLayout(self.help_widget)
        self.tabs.addTab(self.help_widget, "Ajuda")

        help_text = QLabel(
            "<h2>Ajuda</h2>"
            "<p><b>Fórmulas disponíveis:</b></p>"
            "<ul>"
            "<li><code>=SUM(A1:A100)</code>: Soma os valores especificados em um intervalo de células.</li>"
            "<li><code>=AVERAGE(A1:A100)</code>: Média dos valores de um intervalo de células.</li>"
            "<li><code>=MAX(A1:A100)</code>: Retorna o valor máximo de um intervalo de células.</li>"
            "<li><code>=MIN(A1:A100)</code>: Retorna o valor mínimo de um intervalo de células.</li>"
            "<li><code>=SUB(A1,A2)</code>: Subtrai os valores das células especificadas.</li>"
            "<li><code>=MUL(A1,A2)</code>: Multiplica os valores das células especificadas.</li>"
            "<li><code>=DIV(A1,A2)</code>: Divide os valores das células especificadas.</li>"
            "<li><code>=CONCATENATE(A1,B1)</code>: Concatena os valores das células especificadas.</li>"
            "</ul>"
            "<p><b>Como usar:</b> Digite diretamente na célula uma fórmula iniciada com '=' para calcular automaticamente. Quando não estiver em edição, a célula exibirá o resultado da fórmula. Para editar a fórmula, clique duas vezes na célula.</p>"
            "<p><b>Formatar Célula:</b> Use os botões 'Cor de Fundo', 'Fonte' ou 'Cor do Texto' para alterar a formatação das células selecionadas.</p>"
            "<p><b>Deletar Célula:</b> Selecione uma ou mais células e clique no botão 'Deletar Célula' ou pressione a tecla Delete para limpar o conteúdo e a formatação.</p>"
        )
        help_text.setWordWrap(True)
        self.help_layout.addWidget(help_text)

    def add_row(self):
        self.table.insertRow(self.table.rowCount())
        self.table.resizeRowsToContents()  # Ajusta as linhas após adicionar uma linha
        # Inicializar a nova célula com formatação padrão
        for col in range(self.table.columnCount()):
            item = QTableWidgetItem("")
            item.setBackground(QColor("#FFFFFF"))  # Fundo branco
            item.setForeground(QColor("#000000"))  # Texto preto
            font = QFont("Arial", 10)
            item.setFont(font)
            self.table.setItem(self.table.rowCount() - 1, col, item)

    def remove_row(self):
        if self.table.rowCount() > 0:
            self.table.removeRow(self.table.rowCount() - 1)
            self.table.resizeRowsToContents()  # Ajusta as linhas após remover uma linha

    def add_column(self):
        col_count = self.table.columnCount()
        self.table.insertColumn(col_count)
        header_label = ""
        if col_count < 26:
            header_label = chr(65 + col_count)
        else:
            # Suporte para mais de 26 colunas (AA, AB, etc.)
            first = (col_count // 26) - 1
            second = col_count % 26
            header_label = chr(65 + first) + chr(65 + second)
        self.table.setHorizontalHeaderItem(col_count, QTableWidgetItem(header_label))
        # Inicializar a nova célula com formatação padrão
        for row in range(self.table.rowCount()):
            item = QTableWidgetItem("")
            item.setBackground(QColor("#FFFFFF"))  # Fundo branco
            item.setForeground(QColor("#000000"))  # Texto preto
            font = QFont("Arial", 10)
            item.setFont(font)
            self.table.setItem(row, col_count, item)

    def remove_column(self):
        if self.table.columnCount() > 0:
            self.table.removeColumn(self.table.columnCount() - 1)

    def has_circular_dependency(self, start_cell, target_cell, visited=None):
        """
        Verifica se a célula `start_cell` depende (direta ou indiretamente) de `target_cell`.
        """
        if visited is None:
            visited = set()
        if start_cell == target_cell:
            return True
        if start_cell in visited:
            return False
        visited.add(start_cell)
        if start_cell in self.dependencies:
            for dep in self.dependencies[start_cell]:
                if self.has_circular_dependency(dep, target_cell, visited):
                    return True
        return False

    
    
    def handle_cell_edit(self, item):
        if self.currently_editing:
            # Se estamos no meio de uma edição iniciada pelo handle_cell_double_clicked, ignore
            self.currently_editing = False
            return

        row, col = item.row(), item.column()
        text = item.text()
        cell_key = (row, col)

        # Remover esta célula das dependências anteriores
        if cell_key in self.dependencies:
            for dep in self.dependencies[cell_key]:
                if dep in self.dependents:
                    self.dependents[dep].discard(cell_key)
                    if not self.dependents[dep]:
                        del self.dependents[dep]
            del self.dependencies[cell_key]

        if text.startswith("="):
            # Parsear a fórmula para encontrar referências
            referenced_cells = self.parse_formula_references(text, current_cell=cell_key)

            # Detectar referências circulares
            circular = False
            for dep in referenced_cells:
                if self.has_circular_dependency(dep, cell_key):
                    circular = True
                    break

            if circular:
                # Definir a célula como erro devido a referência circular
                self.table.blockSignals(True)
                item.setText("#CIRCULAR")
                self.table.blockSignals(False)
                # Não atualizar dependências
                return

            # Armazenar a fórmula
            item.setData(Qt.UserRole, text)
            self.dependencies[cell_key] = referenced_cells
            for dep in referenced_cells:
                if dep not in self.dependents:
                    self.dependents[dep] = set()
                self.dependents[dep].add(cell_key)

            # Avaliar a fórmula
            result = self.evaluate_formula(text)
            self.table.blockSignals(True)
            item.setText(self.format_number(result))
            self.table.blockSignals(False)

            # Atualizar dependentes
            self.update_dependents(cell_key)

            # Atualizar a barra de fórmulas se a célula estiver selecionada
            if self.is_cell_selected(row, col):
                self.formula_bar.setText(text)
        else:
            # Remover qualquer fórmula armazenada
            item.setData(Qt.UserRole, None)
            # Atualizar dependentes
            self.update_dependents(cell_key)
            # Atualizar a barra de fórmulas se a célula estiver selecionada
            if self.is_cell_selected(row, col):
                self.formula_bar.setText(text)

        # Ajustar as colunas e linhas após a edição
        self.table.resizeColumnsToContents()
        self.table.resizeRowsToContents()  # Ajusta as linhas após a edição




    def parse_formula_references(self, formula, current_cell=None):
        """
        Analisa a fórmula e retorna um conjunto de células referenciadas.
        Exclui a própria célula sendo editada para evitar detecção incorreta de ciclos.
        """
        # Remover o sinal de igual
        expr = formula[1:]
        # Encontrar intervalos e células individuais
        range_pattern = r'([A-Z]+[1-9][0-9]*):([A-Z]+[1-9][0-9]*)'
        cell_pattern = r'([A-Z]+[1-9][0-9]*)'

        references = set()

        # Encontrar intervalos
        for match in re.finditer(range_pattern, expr):
            start_ref = match.group(1)
            end_ref = match.group(2)
            start_row, start_col = self.parse_cell_reference(start_ref)
            end_row, end_col = self.parse_cell_reference(end_ref)
            for r in range(min(start_row, end_row), max(start_row, end_row) + 1):
                for c in range(min(start_col, end_col), max(start_col, end_col) + 1):
                    ref = (r, c)
                    if ref != current_cell:
                        references.add(ref)

        # Encontrar células individuais
        for match in re.finditer(cell_pattern, expr):
            cell_ref = match.group(1)
            row, col = self.parse_cell_reference(cell_ref)
            ref = (row, col)
            if ref != current_cell:
                references.add(ref)

        return references

    def update_dependents(self, cell_key, visited=None):
        if visited is None:
            visited = set()
        if cell_key in visited:
            return  # Já processado
        visited.add(cell_key)

        if cell_key in self.dependents:
            for dependent in self.dependents[cell_key]:
                dependent_item = self.table.item(dependent[0], dependent[1])
                if dependent_item:
                    formula = dependent_item.data(Qt.UserRole)
                    if formula:
                        result = self.evaluate_formula(formula)
                        if result in ("#CIRCULAR", "#ERRO"):
                            # Definir a célula como erro sem exibir pop-up
                            self.table.blockSignals(True)
                            dependent_item.setText(result)
                            self.table.blockSignals(False)
                        else:
                            self.table.blockSignals(True)
                            dependent_item.setText(self.format_number(result))
                            self.table.blockSignals(False)
                            # Atualizar dependentes recursivamente
                            self.update_dependents(dependent, visited)

    
    def format_number(self, value):
        try:
            num = float(value)
            if num.is_integer():
                return str(int(num))
            else:
                return str(num)
        except (ValueError, TypeError):
            return str(value)  # Garantir que sempre retorne uma string


    def update_formula_bar(self):
        selected_items = self.table.selectedItems()
        if selected_items:
            item = selected_items[0]
            formula = item.data(Qt.UserRole)
            if formula:
                self.formula_bar.setText(formula)
            else:
                cell_text = item.text() if item.text() else ""
                self.formula_bar.setText(cell_text)
        else:
            self.formula_bar.setText("")

    def apply_formula(self):
        # Este método pode ser opcional agora, já que estamos permitindo digitar diretamente nas células
        pass

    def evaluate_formula(self, formula):
        # Remover o sinal de igual e avaliar a expressão
        expr = formula[1:]
        try:
            # Avaliar a expressão de forma segura
            # Limitar o escopo do eval para evitar execuções perigosas
            allowed_names = {
                'SUM': self.sum_func,
                'AVERAGE': self.average_func,
                'MAX': max,
                'MIN': min,
                'SUB': self.sub_func,
                'MUL': self.mul_func,
                'DIV': self.div_func,
                'CONCATENATE': self.concatenate_func
            }
            # Substituir referências de células e intervalos por valores
            expr = self.replace_cell_references(expr)
            # Avaliar a expressão
            result = eval(expr, {"__builtins__": None}, allowed_names)
            return result
        except Exception:
            return "#ERRO"

    def replace_cell_references(self, expr):
        # Primeiro, tratar intervalos como A1:A5
        range_pattern = r'([A-Z]+[1-9][0-9]*):([A-Z]+[1-9][0-9]*)'

        def range_replacer(match):
            start_ref = match.group(1)
            end_ref = match.group(2)
            start_row, start_col = self.parse_cell_reference(start_ref)
            end_row, end_col = self.parse_cell_reference(end_ref)
            values = []
            for r in range(min(start_row, end_row), max(start_row, end_row) + 1):
                for c in range(min(start_col, end_col), max(start_col, end_col) + 1):
                    item = self.table.item(r, c)
                    if item:
                        value = item.text()
                        if value.startswith("="):
                            # Recursivamente avaliar a fórmula
                            value = str(self.evaluate_formula(value))
                        try:
                            num = float(value)
                        except:
                            num = 0
                        values.append(num)
                    else:
                        values.append(0)
            return str(values)  # Substitui o intervalo por uma lista de valores

        expr = re.sub(range_pattern, range_replacer, expr)

        # Depois, tratar referências individuais como A1, B2, etc.
        cell_pattern = r'([A-Z]+[1-9][0-9]*)'

        def cell_replacer(match):
            cell_ref = match.group(1)
            row, col = self.parse_cell_reference(cell_ref)
            item = self.table.item(row, col)
            if item:
                value = item.text()
                if value.startswith("="):
                    # Recursivamente avaliar a fórmula
                    return str(self.evaluate_formula(value))
                try:
                    return str(float(value))
                except:
                    return "0"
            return "0"

        expr = re.sub(cell_pattern, cell_replacer, expr)

        return expr

    def sum_func(self, *args):
        if len(args) == 1 and isinstance(args[0], list):
            return sum(args[0])
        return sum(args)

    def average_func(self, *args):
        if len(args) == 1 and isinstance(args[0], list):
            return sum(args[0]) / len(args[0]) if args[0] else 0
        return sum(args) / len(args) if args else 0

    def sub_func(self, *args):
        if not args:
            return "#ERRO"
        result = args[0]
        for num in args[1:]:
            result -= num
        return result

    def mul_func(self, *args):
        result = 1
        for num in args:
            result *= num
        return result

    def div_func(self, *args):
        if not args:
            return "#ERRO"
        result = args[0]
        try:
            for num in args[1:]:
                if num == 0:
                    return "#ERRO"
                result /= num
            return result
        except ZeroDivisionError:
            return "#ERRO"

    def concatenate_func(self, *args):
        return ''.join(map(str, args))

    def parse_cell_reference(self, ref):
        match = ref.strip().upper()
        # Separar letras e números
        letters = re.match(r'[A-Z]+', match).group()
        numbers = re.search(r'[0-9]+', match).group()
        # Converter letras para índice de coluna (A=0, B=1, ..., Z=25, AA=26, etc.)
        col = 0
        for char in letters:
            col = col * 26 + (ord(char) - ord('A') + 1)
        col -= 1  # Índice baseado em 0
        row = int(numbers) - 1  # Índice baseado em 0
        return row, col

    
    def handle_cell_double_clicked(self, row, col):
        item = self.table.item(row, col)
        if item:
            formula = item.data(Qt.UserRole)
            if formula:
                self.table.blockSignals(True)          # Bloquear sinais
                item.setText(formula)                  # Mostrar a fórmula para edição
                self.table.blockSignals(False)         # Desbloquear sinais
                self.table.editItem(item)
 
     

    def is_cell_selected(self, row, col):
        selected = self.table.selectedItems()
        for item in selected:
            if item.row() == row and item.column() == col:
                return True
        return False

    def delete_cell_button(self):
        self.delete_cell()

   
    def delete_cell(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Aviso", "Por favor, selecione uma célula para deletar.")
            return

        for item in selected_items:
            row, col = item.row(), item.column()
            cell_key = (row, col)
            # Remover a fórmula se existir
            item.setData(Qt.UserRole, None)
            # Remover dependências antigas
            if cell_key in self.dependencies:
                for dep in self.dependencies[cell_key]:
                    if dep in self.dependents:
                        self.dependents[dep].discard(cell_key)
                        if not self.dependents[dep]:
                            del self.dependents[dep]
                del self.dependencies[cell_key]
            # Remover esta célula dos dependentes das dependências anteriores
            if cell_key in self.dependents:
                for dependent in self.dependents[cell_key]:
                    dependent_item = self.table.item(dependent[0], dependent[1])
                    if dependent_item:
                        formula = dependent_item.data(Qt.UserRole)
                        if formula:
                            result = self.evaluate_formula(formula)
                            self.table.blockSignals(True)
                            dependent_item.setText(self.format_number(result))  # <-- Correção aqui
                            self.table.blockSignals(False)
                del self.dependents[cell_key]
            self.table.blockSignals(True)
            item.setText("")  # Limpa o texto
            item.setBackground(QColor("#FFFFFF"))  # Reseta a cor de fundo
            item.setForeground(QColor("#000000"))  # Reseta a cor da fonte para preto
            font = QFont()
            font.setFamily("Arial")  # Define uma família padrão
            font.setPointSize(10)     # Define um tamanho padrão
            item.setFont(font)        # Reseta a fonte para padrão
            self.table.blockSignals(False)
            # Reavaliar dependentes
            self.update_dependents(cell_key)



    def save_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Salvar Planilha", "planilha.xlsx", "Arquivos Excel (*.xlsx)")
        if path:
            if not path.endswith(".xlsx"):
                path += ".xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Planilha"

            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item and (item.text() or item.data(Qt.UserRole)):
                        cell = ws.cell(row=row+1, column=col+1, value=item.text())

                        # Aplicar formatação de fonte
                        font = item.font()
                        font_color = item.foreground().color().name() if item.foreground().color().isValid() else "#000000"
                        openpyxl_font = OpenpyxlFont(
                            name=font.family(),
                            size=font.pointSize(),
                            bold=font.bold,
                            italic=font.italic,
                            underline='single' if font.underline() else 'none',
                            color=font_color.replace("#", "")
                        )
                        cell.font = openpyxl_font

                        # Aplicar cor de fundo
                        bg_color = item.background().color().name() if item.background().color().isValid() else "#FFFFFF"
                        openpyxl_fill = PatternFill(
                            start_color=bg_color.replace("#", ""),
                            end_color=bg_color.replace("#", ""),
                            fill_type="solid"
                        )
                        cell.fill = openpyxl_fill

                        # Aplicar fórmula se existir
                        formula = item.data(Qt.UserRole)
                        if formula:
                            cell.value = formula

            try:
                wb.save(path)
                QMessageBox.information(self, "Sucesso", "Planilha salva com sucesso!")
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Falha ao salvar a planilha:\n{str(e)}")

    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Carregar Planilha", "", "Arquivos Excel (*.xlsx)")
        if path:
            try:
                wb = openpyxl.load_workbook(path)
                ws = wb.active

                rows = ws.max_row
                columns = ws.max_column
                self.table.blockSignals(True)
                self.table.setRowCount(rows)
                self.table.setColumnCount(columns)
                self.table.setHorizontalHeaderLabels([chr(65 + i) for i in range(columns)])

                # Limpar dependências existentes
                self.dependencies.clear()
                self.dependents.clear()

                for row in range(1, rows + 1):
                    for col in range(1, columns + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            item = QTableWidgetItem(str(cell.value))

                            # Aplicar formatação de fundo
                            bg_color = cell.fill.start_color.index if cell.fill.start_color.type == "rgb" else "#FFFFFF"
                            if bg_color and bg_color != "00000000":
                                # Extraindo os últimos 6 caracteres para obter a cor RGB
                                color_hex = f"#{bg_color[-6:]}" if len(bg_color) >= 6 else "#FFFFFF"
                                item.setBackground(QColor(color_hex))
                            else:
                                item.setBackground(QColor("#FFFFFF"))  # Garantir fundo branco

                            # Aplicar formatação de fonte
                            font = QFont()
                            font.setFamily(cell.font.name if cell.font.name else "Arial")
                            font_size = cell.font.size if cell.font.size else 10
                            if isinstance(font_size, float):
                                font.setPointSize(int(font_size))
                            else:
                                font.setPointSize(font_size)
                            font.setBold(bool(cell.font.bold))       # Correção aplicada aqui
                            font.setItalic(bool(cell.font.italic))   # Correção aplicada aqui
                            font.setUnderline(bool(cell.font.underline))  # Correção aplicada aqui
                            item.setFont(font)

                            # Aplicar cor da fonte
                            font_color = cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else "#000000"
                            if font_color and len(font_color) >= 6:
                                item.setForeground(QColor(f"#{font_color[-6:]}"))
                            else:
                                item.setForeground(QColor("#000000"))  # Garantir texto preto

                            # Verificar se há fórmula
                            if cell.data_type == 'f':
                                item.setData(Qt.UserRole, cell.value)
                                # Atualizar dependências
                                referenced_cells = self.parse_formula_references(cell.value, current_cell=(row-1, col-1))
                                self.dependencies[(row-1, col-1)] = referenced_cells
                                for ref in referenced_cells:
                                    if ref not in self.dependents:
                                        self.dependents[ref] = set()
                                    self.dependents[ref].add((row-1, col-1))
                                # Avaliar a fórmula
                                result = self.evaluate_formula(cell.value)
                                item.setText(self.format_number(result))
                            else:
                                item.setData(Qt.UserRole, None)

                            self.table.setItem(row-1, col-1, item)

                # Garantir que todas as células sem dados explicitamente salvos tenham formatação padrão
                for row in range(rows):
                    for col in range(columns):
                        item = self.table.item(row, col)
                        if not item:
                            item = QTableWidgetItem("")
                            item.setBackground(QColor("#FFFFFF"))
                            item.setForeground(QColor("#000000"))
                            font = QFont("Arial", 10)
                            item.setFont(font)
                            self.table.setItem(row, col, item)

                self.table.blockSignals(False)
                self.table.resizeColumnsToContents()
                self.table.resizeRowsToContents()
                QMessageBox.information(self, "Sucesso", "Planilha carregada com sucesso!")
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Falha ao carregar a planilha:\n{str(e)}")

    def format_background(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Aviso", "Por favor, selecione uma célula para formatar o fundo.")
            return

        color = QColorDialog.getColor()
        if color.isValid():
            self.table.blockSignals(True)  # Bloquear sinais antes da formatação
            for item in selected_items:
                item.setBackground(color)  # Aplica a cor de fundo
            self.table.blockSignals(False)  # Desbloquear sinais após a formatação

    def format_font(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Aviso", "Por favor, selecione uma célula para formatar a fonte.")
            return

        font, ok = QFontDialog.getFont()
        if ok:
            self.table.blockSignals(True)  # Bloquear sinais antes da formatação
            for item in selected_items:
                item.setFont(font)  # Aplica a fonte selecionada
            self.table.blockSignals(False)  # Desbloquear sinais após a formatação

    def format_text_color(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Aviso", "Por favor, selecione uma célula para alterar a cor do texto.")
            return

        color = QColorDialog.getColor()
        if color.isValid():
            self.table.blockSignals(True)  # Bloquear sinais antes da formatação
            for item in selected_items:
                item.setForeground(color)  # Aplica a cor do texto
            self.table.blockSignals(False)  # Desbloquear sinais após a formatação


def main():
    app = QApplication(sys.argv)

    # Definir o estilo para evitar aviso de estilo inválido
    app.setStyle("Fusion")

    window = SimpleExcelWidget()
    window.setWindowTitle("Simple Excel")
    window.resize(1000, 600)
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()

